import sys
import re
from datetime import datetime
from pathlib import Path
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, QTimer, QUrl
from PyQt5.QtGui import QFont, QIcon, QPixmap, QPainter, QColor, QPen, QBrush, QRadialGradient

from app.core.data_loader import DATA_LOADER
from app.core.quiz_manager import QuizManager
from app.core.search_engine import SearchEngine
from app.core.update_checker import UpdateChecker
from app.ui.top_bar import TopBar
from app.ui.left_panel import LeftPanel
from app.ui.right_view import RightViewPanel
from app.ui.right_quiz import RightQuizPanel
from app.ui_components.pdf_viewer import PDFViewerDialog
from app.ui_components.login_dialog import LoginDialog
from app.core.license_validator import LicenseValidator


class MainWindow(QMainWindow):
    def __init__(self, software_version="2.0.2", parent=None):
        super().__init__(parent)
        self.software_version = software_version
        self.setWindowTitle("工位SOP助手 - 智能工站系统")
        self.setWindowIcon(self.create_icon())
        self.setGeometry(120, 80, 1280, 820)

        # 核心数据
        self.knowledge_data = DATA_LOADER.get_knowledge_graph()
        self.station_to_pdfs = DATA_LOADER.get_station_to_pdfs()
        self.file_categories = DATA_LOADER.get_file_categories()
        self.all_stations = DATA_LOADER.get_all_stations()

        docs = self.knowledge_data.get("documents", [])
        self.pdf_to_doc = {doc.get("pdf_name", ""): True for doc in docs if doc.get("pdf_name")}

        self.current_station = self.all_stations[0] if self.all_stations else "无工站"
        self.logged_in = False
        self.current_user = ""
        self._current_selected_pdf = None

        # 考核相关
        self.quiz_manager = QuizManager(self)
        self.quiz_questions = []
        self.quiz_answer_widgets = []
        self.quiz_pdf_name = None
        self.quiz_version = None
        self.quiz_records = self.quiz_manager.quiz_records

        # 搜索相关
        self.search_engine = SearchEngine(self)
        self.search_index = []
        self.current_results = []

        # 更新检查（传入版本号）
        self.update_checker = UpdateChecker(self, self.software_version)

        # UI组件
        self.top_bar = None
        self.left_panel = None
        self.right_view = None
        self.right_quiz = None
        self.splitter = None

        self.init_ui()
        self.refresh_left_list()
        self.update_right_panel()

        QTimer.singleShot(200, self.sync_top_layout)

        # 定时器：每小时静默同步（使用 update_checker）
        self.update_timer = QTimer(self)
        self.update_timer.timeout.connect(self._on_silent_sync)
        self.update_timer.start(3600 * 1000)  # 1小时

    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)

        self.setMinimumSize(800, 600)

        # 顶部工具栏
        self.top_bar = TopBar(self)
        self.top_bar.setFixedHeight(144)
        self.top_bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.top_bar.setStyleSheet("""
            QWidget {
                background-color: #f5f7fa;
                border-radius: 6px;
                border: 1px solid #d0d7de;
            }
        """)
        main_layout.addWidget(self.top_bar)
        self.top_bar.set_stations(self.all_stations)

        # 中间分割
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.setChildrenCollapsible(False)

        self.left_panel = LeftPanel(self)
        self.splitter.addWidget(self.left_panel)

        self.right_stack = QStackedWidget()
        self.right_view = RightViewPanel(self)
        self.right_quiz = RightQuizPanel(self)
        self.right_stack.addWidget(self.right_view)
        self.right_stack.addWidget(self.right_quiz)
        self.splitter.addWidget(self.right_stack)

        self.splitter.setStretchFactor(0, 2)
        self.splitter.setStretchFactor(1, 1)
        self.splitter.splitterMoved.connect(self.sync_top_layout)
        main_layout.addWidget(self.splitter)

        # 状态栏（使用 statusBar() 方法）
        self.statusBar().showMessage("就绪 | 请先登录")

        # 版本信息
        version_widget = QWidget()
        version_layout = QHBoxLayout(version_widget)
        version_layout.setContentsMargins(0, 0, 0, 0)
        version_layout.setSpacing(4)
        self.version_label = QLabel()
        self.version_label.setStyleSheet("color: #888;")
        self.update_version_label()
        version_layout.addWidget(self.version_label)

        self.refresh_btn = QPushButton("↻")
        self.refresh_btn.setFixedSize(24, 24)
        self.refresh_btn.setToolTip("检查并同步云端数据更新")
        self.refresh_btn.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: none;
                font-size: 18px;
                color: #888;
            }
            QPushButton:hover {
                color: #3498db;
            }
        """)
        # 绑定手动刷新（非静默，带进度）
        self.refresh_btn.clicked.connect(self._on_refresh_clicked)
        version_layout.addWidget(self.refresh_btn)

        self.statusBar().addPermanentWidget(version_widget)

    # ---------- 同步与调整 ----------
    def sync_top_layout(self):
        if not hasattr(self, 'splitter'):
            return
        left_width = self.splitter.widget(0).width()
        right_width = self.splitter.widget(1).width()
        self.top_bar.left_area.setFixedWidth(left_width)
        self.top_bar.search_area.setFixedWidth(right_width)
        self.adjust_search_width()

    def adjust_search_width(self):
        if not self.top_bar.search_input:
            return
        container_width = self.top_bar.search_area.width()
        if container_width <= 0:
            return
        target = int(container_width * 0.8)
        self.top_bar.search_input.setMinimumWidth(target)
        self.top_bar.search_input.setMaximumWidth(target)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        QTimer.singleShot(50, self.sync_top_layout)

    # ---------- 登录/退出 ----------
    def toggle_login(self):
        if self.logged_in:
            # 退出登录
            self.logged_in = False
            self.current_user = ""
            self.top_bar.user_label.setText("👤 未登录")
            self.top_bar.user_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12pt;")
            self.top_bar.login_btn.setText("登录")
            if self.top_bar.exam_toggle.isChecked():
                self.top_bar.exam_toggle.setChecked(False)
            # 重置工站为第一个
            if self.all_stations:
                self.current_station = self.all_stations[0]
            else:
                self.current_station = "无工站"
            self.refresh_left_list()
            self.update_right_panel()
            self.statusBar().showMessage("已退出")
        else:
            dialog = LoginDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                username = dialog.get_username()
                if username:
                    self.logged_in = True
                    self.current_user = username
                    self.top_bar.user_label.setText(f"👤 {username} (已登录)")
                    self.top_bar.user_label.setStyleSheet("font-weight: bold; color: #2c3e50; font-size: 12pt;")
                    self.top_bar.login_btn.setText("退出")
                    # 同步当前工站为下拉框选中的值
                    self.current_station = self.top_bar.station_combo.currentText()
                    # 调用工站切换方法以刷新界面
                    self.on_station_changed(self.current_station)
                    self.statusBar().showMessage(f"欢迎, {username}")
                    if self.top_bar.exam_toggle.isChecked():
                        self.top_bar.exam_toggle.setChecked(False)


    # ---------- 工站/分类 ----------
    def on_station_changed(self, station_name):
        if not self.logged_in:
            self.statusBar().showMessage("请先登录")
            return
        self.current_station = station_name
        self.top_bar.search_input.blockSignals(True)
        self.top_bar.search_input.clear()
        self.top_bar.search_input.blockSignals(False)
        self.refresh_left_list()
        self.update_right_panel()
        self.statusBar().showMessage(f"切换到工站: {station_name}")

    def on_category_filter_changed(self):
        if not self.logged_in:
            return
        self.refresh_left_list()
        self.top_bar.search_input.blockSignals(True)
        self.top_bar.search_input.clear()
        self.top_bar.search_input.blockSignals(False)
        self.update_right_panel()

    # ---------- 左面板 ----------
    def refresh_left_list(self):
        self.left_panel.clear()
        if not self.logged_in:
            self.left_panel.set_placeholder("请先登录以查看文件")
            return

        pdf_list = self.station_to_pdfs.get(self.current_station, [])
        show_general = self.top_bar.checkbox_general.isChecked()
        show_product = self.top_bar.checkbox_product.isChecked()
        if show_general or show_product:
            filtered = []
            for pdf_name in pdf_list:
                category = self.file_categories.get(pdf_name, "")
                if show_general and category == "通用操作":
                    filtered.append(pdf_name)
                elif show_product and category == "产品相关":
                    filtered.append(pdf_name)
            pdf_list = filtered

        search_text = self.top_bar.search_input.text().strip()
        if search_text:
            pdf_list = [p for p in pdf_list if search_text.lower() in p.lower()]

        is_quiz_mode = self.top_bar.exam_toggle.isChecked()
        available_quiz_names = DATA_LOADER.get_available_quiz_names() if is_quiz_mode else []

        for pdf_name in pdf_list:
            if is_quiz_mode:
                if pdf_name not in available_quiz_names:
                    continue

            if is_quiz_mode:
                record = self.quiz_records.get(pdf_name, {})
                score = record.get("score", 0)
                if score > 0:
                    color = QuizManager.score_to_color(score)
                    self.left_panel.add_item(f"{pdf_name}  {score}分", QColor(color[0], color[1], color[2]))
                else:
                    self.left_panel.add_item(pdf_name, Qt.red)
            else:
                if pdf_name in self.pdf_to_doc:
                    self.left_panel.add_item(pdf_name, Qt.green)
                else:
                    self.left_panel.add_item(pdf_name, Qt.red)

        if self.left_panel.count() == 0:
            self.left_panel.set_placeholder("(暂无匹配的SOP文件)")

        self.search_engine.build_dynamic_index()

    # ---------- 右面板 ----------
    def update_right_panel(self):
        if not self.logged_in:
            self.right_view.set_html("<p style='color:#7f8c8d; font-size: 14pt;'>请先登录以查看工站更新推送。</p>")
            return

        pdf_list = self.station_to_pdfs.get(self.current_station, [])
        show_general = self.top_bar.checkbox_general.isChecked()
        show_product = self.top_bar.checkbox_product.isChecked()
        if show_general or show_product:
            filtered = []
            for pdf_name in pdf_list:
                category = self.file_categories.get(pdf_name, "")
                if show_general and category == "通用操作":
                    filtered.append(pdf_name)
                elif show_product and category == "产品相关":
                    filtered.append(pdf_name)
            pdf_list = filtered

        all_docs = self.knowledge_data.get("documents", [])
        doc_map = {doc.get("pdf_name", ""): doc for doc in all_docs}
        updates = []
        for pdf_name in pdf_list:
            doc = doc_map.get(pdf_name)
            if not doc:
                continue
            history_node = None
            history_page = "1"
            for proc in doc.get("processes", []):
                if proc.get("name") == "履历":
                    history_node = proc
                    history_page = proc.get("page", "1")
                    break
            if not history_node:
                continue
            history_list = history_node.get("content", [])
            if not history_list:
                continue
            latest = history_list[-1] if history_list else None
            if not latest:
                continue
            doc_name = doc.get("doc_name", pdf_name)
            revision_date = latest.get("revision_date", "")
            revision_content = latest.get("revision_content", "")
            version = latest.get("version", "")
            if len(revision_content) > 80:
                revision_content = revision_content[:80] + "..."
            updates.append({
                "pdf_name": pdf_name,
                "doc_name": doc_name,
                "date": revision_date,
                "content": revision_content,
                "version": version,
                "page": history_page
            })

        if not updates:
            self.right_view.set_html(f"<p style='color:#7f8c8d; font-size: 14pt;'>当前工站 <b>{self.current_station}</b> 暂无文档更新记录。</p>")
            return

        # 日期解析辅助函数
        from datetime import datetime
        def parse_date(date_str):
            if not date_str:
                return datetime.min
            for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            return datetime.min

        updates.sort(key=lambda x: parse_date(x['date']), reverse=True)

        html_parts = [f"<div style='font-family: Microsoft YaHei, sans-serif;'>",
                    f"<h2 style='color: #2c3e50; margin-bottom: 15px;'>📢 {self.current_station} 动态推送</h2>",
                    "<div style='display: flex; flex-direction: column; gap: 14px;'>"]
        for update in updates:
            pdf_name = update['pdf_name']
            page = update['page']
            highlight = update['date']
            link = f'openpdf://open?file={pdf_name}&page={page}&highlight={highlight}'
            html_parts.append(f"""
            <div style='background: #ffffff; border-left: 4px solid #3498db; border-radius: 8px; padding: 14px 18px; box-shadow: 0 4px 12px rgba(0,0,0,0.12); transition: box-shadow 0.2s; margin-bottom: 4px;'>
                <div style='display: flex; justify-content: space-between; align-items: center;'>
                    <span style='font-size: 14pt; font-weight: bold; color: #1a3c6e;'>
                        📄 {update['doc_name']}
                        <span style='font-size: 12pt; font-weight: normal; color: #888; margin-left: 8px;'>
                            [{update.get('version', '')}]
                        </span>
                    </span>
                    <a href='{link}' style='color: #0066cc; text-decoration: underline; font-size: 11pt; cursor: pointer; background: transparent; padding: 0;'>跳转</a>
                </div>
                <div style='margin-top: 6px; font-size: 13pt; color: #555;'>📅 {update['date']}</div>
                <div style='margin-top: 4px; font-size: 13pt; color: #333; line-height: 1.5;'>{update['content']}</div>
            </div>
            """)
        html_parts.append("</div></div>")
        self.right_view.set_html("".join(html_parts))

    # ---------- 考核模式 ----------
    def on_exam_mode_toggled(self, checked):
        if checked and not self.logged_in:
            QMessageBox.warning(self, "提示", "请先登录才能进入考核模式")
            self.top_bar.exam_toggle.setChecked(False)
            return
        if checked:
            self.top_bar.exam_toggle.setText("📝 查看模式")
            self.right_stack.setCurrentIndex(1)
            self.top_bar.set_search_visible(False)
            self.statusBar().showMessage("已进入考核模式")
            self.quiz_manager.load_records()
            self.refresh_left_list()
            if self._current_selected_pdf:
                self.load_quiz_for_pdf(self._current_selected_pdf)
            else:
                self.right_quiz.clear_quiz()
                label = QLabel("请从左侧选择一个文件以加载题库。")
                label.setStyleSheet("color: #7f8c8d; font-size: 14px; padding: 16px;")
                self.right_quiz.container_layout.insertWidget(0, label)
                self.right_quiz.enable_submit(False)
        else:
            self.top_bar.exam_toggle.setText("📝 考核模式")
            self.right_stack.setCurrentIndex(0)
            self.top_bar.set_search_visible(True)
            self.statusBar().showMessage("已退出考核模式")
            self.quiz_manager.save_records()
            self.update_right_panel()
            self.refresh_left_list()

    def load_quiz_for_pdf(self, pdf_name):
        self.right_quiz.clear_quiz()
        self.right_quiz.enable_buttons(False)
        self.right_quiz.enable_submit(False)

        normalized_name = pdf_name.strip()
        quiz_bank = DATA_LOADER.get_quiz_bank()
        quiz_info = quiz_bank.get(normalized_name)
        if quiz_info is None:
            for key in quiz_bank:
                if key.lower() == normalized_name.lower():
                    quiz_info = quiz_bank[key]
                    normalized_name = key
                    break

        if quiz_info is None:
            label = QLabel("该文档暂无题库。")
            label.setStyleSheet("color: #7f8c8d; font-size: 16px; padding: 16px;")
            self.right_quiz.container_layout.insertWidget(0, label)
            self.right_quiz.enable_submit(False)
            return

        if isinstance(quiz_info, list):
            questions = quiz_info
            version = "未知版本"
        else:
            questions = quiz_info.get("questions", [])
            version = quiz_info.get("version", "未知版本")

        if not questions:
            label = QLabel("该文档题库为空。")
            label.setStyleSheet("color: #7f8c8d; font-size: 16px; padding: 16px;")
            self.right_quiz.container_layout.insertWidget(0, label)
            self.right_quiz.enable_submit(False)
            return

        self.quiz_questions = questions
        self.quiz_pdf_name = normalized_name
        self.quiz_version = version

        record = self.quiz_records.get(normalized_name, {})
        score = record.get("score", 0)
        submitted = record.get("submitted", False)
        saved_answers = record.get("answers", [])

        if score > 0:
            color = QuizManager.score_to_color(score)
            self.right_quiz.set_score(score, color)
        else:
            self.right_quiz.set_score(0, None)

        self.render_quiz(questions, submitted=submitted, saved_answers=saved_answers)

        if saved_answers and len(saved_answers) == len(self.quiz_answer_widgets):
            for idx, ans in enumerate(saved_answers):
                if not ans:
                    continue
                item = self.quiz_answer_widgets[idx]
                if item['type'] == 'single':
                    for i, btn in enumerate(item['widgets']):
                        if chr(65 + i) == ans:
                            btn.setChecked(True)
                            break
                else:
                    for i, btn in enumerate(item['widgets']):
                        btn.setChecked(chr(65 + i) in ans)

        self.right_quiz.enable_submit(True)
        self.right_quiz.enable_buttons(True)

    def render_quiz(self, questions, submitted=False, saved_answers=None):
        from PyQt5.QtWidgets import QGroupBox, QRadioButton, QCheckBox, QButtonGroup, QVBoxLayout, QLabel

        self.quiz_answer_widgets.clear()

        option_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
        for idx, q in enumerate(questions):
            group_box = QGroupBox()
            group_box.setStyleSheet("""
                QGroupBox {
                    border: 1px solid #d0d7de;
                    border-radius: 6px;
                    margin-top: 12px;
                    padding-top: 8px;
                    font-weight: bold;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 12px;
                    padding: 0 8px;
                    color: #1a3c6e;
                    font-size: 15px;
                }
            """)
            title = f"第 {idx+1} 题（{'单选' if q['type'] == 'single' else '多选'}）"
            group_box.setTitle(title)

            vbox = QVBoxLayout(group_box)
            vbox.setSpacing(10)

            question_label = QLabel(q['question'])
            question_label.setWordWrap(True)
            question_label.setStyleSheet("font-size: 24px; padding: 4px 0;")
            vbox.addWidget(question_label)

            option_widgets = []
            if q['type'] == 'single':
                btn_group = QButtonGroup()
                for i, opt_text in enumerate(q['options']):
                    label_text = f"{option_labels[i]}. {opt_text}" if i < len(option_labels) else opt_text
                    radio = QRadioButton(label_text)
                    radio.setStyleSheet("font-size: 22px; padding: 4px 0;")
                    vbox.addWidget(radio)
                    btn_group.addButton(radio, i)
                    option_widgets.append(radio)
                self.quiz_answer_widgets.append({
                    'type': 'single',
                    'group': btn_group,
                    'widgets': option_widgets,
                    'answer': q['answer']
                })
            else:
                for i, opt_text in enumerate(q['options']):
                    label_text = f"{option_labels[i]}. {opt_text}" if i < len(option_labels) else opt_text
                    checkbox = QCheckBox(label_text)
                    checkbox.setStyleSheet("font-size: 22px; padding: 4px 0;")
                    vbox.addWidget(checkbox)
                    option_widgets.append(checkbox)
                self.quiz_answer_widgets.append({
                    'type': 'multi',
                    'widgets': option_widgets,
                    'answer': q['answer']
                })

            if submitted:
                correct_ans = q['answer']
                user_ans = saved_answers[idx] if saved_answers and idx < len(saved_answers) else "未选"
                is_correct = (user_ans == correct_ans)
                result_text = "✅ 正确" if is_correct else "❌ 错误"
                result_color = "#27ae60" if is_correct else "#e74c3c"
                correct_display = f"正确答案：{correct_ans}"
                user_display = f"您的答案：{user_ans if user_ans else '未选'}"
                result_label = QLabel(f"{result_text} | {correct_display} | {user_display}")
                result_label.setStyleSheet(f"font-size: 14px; color: {result_color}; padding: 4px; background: #f9f9f9; border-radius: 4px;")
                vbox.addWidget(result_label)

            self.right_quiz.container_layout.insertWidget(self.right_quiz.container_layout.count() - 1, group_box)

        if questions:
            tip_label = QLabel("请完成所有题目后点击「提交答案」")
            tip_label.setStyleSheet("color: #888; font-size: 14px; padding: 4px;")
            self.right_quiz.container_layout.insertWidget(self.right_quiz.container_layout.count() - 1, tip_label)

    # ---------- 考核按钮功能 ----------
    def on_quiz_open_pdf(self):
        if self.quiz_pdf_name:
            self.open_pdf_by_name(self.quiz_pdf_name, 1, "")
        else:
            QMessageBox.warning(self, "提示", "请先选择一个文档。")

    def on_quiz_restart(self):
        if not self.quiz_pdf_name:
            QMessageBox.warning(self, "提示", "请先选择一个文档。")
            return
        if self.quiz_pdf_name in self.quiz_records:
            self.quiz_records[self.quiz_pdf_name] = {
                "version": self.quiz_version,
                "answers": [],
                "score": 0,
                "submitted": False
            }
            self.quiz_manager.save_records()
        self.load_quiz_for_pdf(self.quiz_pdf_name)
        self.right_quiz.enable_submit(True)
        QMessageBox.information(self, "重新答题", "已重置答题状态。")

    def on_quiz_save_answers(self):
        if not self.quiz_pdf_name:
            QMessageBox.warning(self, "提示", "请先选择一个文档。")
            return

        answers = []
        for item in self.quiz_answer_widgets:
            if item['type'] == 'single':
                selected = item['group'].checkedButton()
                ans = chr(65 + item['group'].buttons().index(selected)) if selected else ""
            else:
                selected_indices = [i for i, cb in enumerate(item['widgets']) if cb.isChecked()]
                ans = ''.join(chr(65 + i) for i in sorted(selected_indices)) if selected_indices else ""
            answers.append(ans)

        record = self.quiz_records.get(self.quiz_pdf_name, {})
        record["version"] = self.quiz_version
        record["answers"] = answers
        self.quiz_records[self.quiz_pdf_name] = record
        self.quiz_manager.save_records()
        QMessageBox.information(self, "保存答案", "答案已保存到记录文件。")

    # ---------- 提交答案 ----------
    def submit_quiz(self):
        import openpyxl
        from datetime import datetime

        if not self.quiz_answer_widgets:
            return

        total = len(self.quiz_answer_widgets)
        correct = 0
        wrong_questions = []

        user_answers = []
        for idx, item in enumerate(self.quiz_answer_widgets):
            if item['type'] == 'single':
                selected = item['group'].checkedButton()
                ans = chr(65 + item['group'].buttons().index(selected)) if selected else ""
            else:
                selected_indices = [i for i, cb in enumerate(item['widgets']) if cb.isChecked()]
                ans = ''.join(chr(65 + i) for i in sorted(selected_indices)) if selected_indices else ""
            user_answers.append(ans)
            if ans == item['answer']:
                correct += 1
            else:
                if ans:
                    wrong_questions.append(str(idx + 1))
                else:
                    wrong_questions.append(str(idx + 1))

        score_percent = int(correct / total * 100) if total > 0 else 0
        score_str = f"{correct}/{total}"

        if self.quiz_pdf_name:
            self.quiz_manager.update_record(
                self.quiz_pdf_name,
                self.quiz_version,
                user_answers,
                score_percent,
                submitted=True
            )
            self.save_quiz_record(self.current_user, self.quiz_pdf_name, self.quiz_version, score_str, wrong_questions)

        color = QuizManager.score_to_color(score_percent)
        self.right_quiz.set_score(score_percent, color)
        self.right_quiz.enable_submit(False)
        self.load_quiz_for_pdf(self.quiz_pdf_name)
        self.refresh_left_list()

    def save_quiz_record(self, user, quiz_name, version, score, wrong_questions):
        import openpyxl
        from datetime import datetime
        from pathlib import Path

        excel_path = Path("assets/personnel_data.xlsx")
        excel_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            if excel_path.exists():
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                headers = ["人员", "题库名", "版本", "得分", "错题", "时间"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            row_num = ws.max_row + 1
            ws.cell(row=row_num, column=1, value=user)
            ws.cell(row=row_num, column=2, value=quiz_name)
            ws.cell(row=row_num, column=3, value=version)
            ws.cell(row=row_num, column=4, value=score)
            ws.cell(row=row_num, column=5, value=", ".join(wrong_questions) if wrong_questions else "无")
            ws.cell(row=row_num, column=6, value=now)

            wb.save(excel_path)
        except Exception as e:
            print(f"保存考核记录失败: {e}")

    # ---------- 搜索 ----------
    def on_search_changed(self, text):
        if not self.logged_in:
            return
        if not text.strip():
            self.update_right_panel()
            return
        if not self.search_index:
            self.search_engine.build_dynamic_index()
        results = self.search_engine.keyword_search(text, self.search_index)
        self.current_results = results
        self.display_search_results(results, text)

    def display_search_results(self, results, query):
        if not results:
            self.right_view.set_html("<p style='color:#7f8c8d; font-size: 12pt;'>😅 未找到相关内容，请尝试其他关键词</p>")
            return

        groups = {}
        for item in results:
            pdf_name = item.get('pdf_name', '')
            key = (item['doc_name'], item['process_name'], item.get('pdf_page', 'N/A'), pdf_name)
            if key not in groups:
                groups[key] = {
                    'doc_name': item['doc_name'],
                    'process_name': item['process_name'],
                    'pdf_page': item.get('pdf_page', 'N/A'),
                    'pdf_name': pdf_name,
                    'items': []
                }
            groups[key]['items'].append({
                'step_item': item.get('step_item', ''),
                'content': item.get('content', ''),
                'precautions': item.get('precautions', '')
            })

        html_parts = [f"<h3>🔍 搜索结果: “{query}”</h3>"]
        html_parts.append(f"<p style='color:#666;'>共 {len(results)} 条匹配记录</p><hr>")
        for idx, (key, group) in enumerate(groups.items(), 1):
            html_parts.append(f"<h4 style='color:#1a3c6e;'>【{idx}】📂 {group['doc_name']}</h4>")
            if group['process_name']:
                html_parts.append(f"<p style='color:#555;'>→ {group['process_name']}</p>")
            page = group['pdf_page']
            pdf_name = group['pdf_name']
            link = f'<a href="openpdf://open?file={pdf_name}&page={page}&highlight={query}">📂 跳转</a>'
            html_parts.append(f"<p style='color:#888;'>📍 页码: {page}  {link}</p>")
            for item in group['items']:
                if item['step_item']:
                    html_parts.append(f"<p style='color:#0b5394; font-weight:bold;'>📌 {item['step_item']}</p>")
                if item['content']:
                    content = item['content']
                    for word in query.split():
                        content = content.replace(word, f"<span style='background-color:#ffeb3b;'>{word}</span>")
                    html_parts.append(f"<p style='margin-left:20px;'>{content}</p>")
                if item['precautions']:
                    html_parts.append(f"<p style='margin-left:20px; color:#d45c00;'>⚠️ 注意事项: {item['precautions']}</p>")
            html_parts.append("<hr>")
        self.right_view.set_html("".join(html_parts))

    # ---------- PDF 打开 ----------
    def open_pdf_by_name(self, pdf_name, page=1, highlight_words=""):
        if not self.logged_in:
            QMessageBox.warning(self, "提示", "请先登录")
            return
        pdf_bytes = DATA_LOADER.get_pdf_bytes(pdf_name)
        if pdf_bytes is None:
            QMessageBox.warning(self, "提示", f"未找到PDF文件: {pdf_name}")
            return
        all_docs = self.knowledge_data.get("documents", [])
        doc_node = None
        for doc in all_docs:
            if doc.get("pdf_name", "") == pdf_name:
                doc_node = doc
                break
        try:
            viewer = PDFViewerDialog(
                pdf_bytes=pdf_bytes,
                pdf_name=pdf_name,
                initial_page=page,
                highlight_words=highlight_words,
                doc_node=doc_node,
                all_docs=all_docs,
                get_pdf_bytes_func=DATA_LOADER.get_pdf_bytes,
                parent=self
            )
            viewer.show()
            self.statusBar().showMessage(f"已打开: {pdf_name} (第{page}页)")
        except Exception as e:
            QMessageBox.critical(self, "PDF查看器错误", f"无法显示PDF：\n{str(e)}")

    def on_anchor_clicked(self, url):
        if url.scheme() == "openpdf":
            query = url.query()
            params = {}
            for pair in query.split("&"):
                if "=" in pair:
                    k, v = pair.split("=", 1)
                    params[k] = v
            pdf_name = params.get("file", "")
            page_str = params.get("page", "1")
            highlight = params.get("highlight", "")
            try:
                page = int(page_str)
            except ValueError:
                page = 1
            if pdf_name:
                self.open_pdf_by_name(pdf_name, page, highlight)
                self.top_bar.search_input.blockSignals(True)
                self.top_bar.search_input.clear()
                self.top_bar.search_input.blockSignals(False)
                self.update_right_panel()

    def on_file_clicked(self, item):
        if not self.logged_in:
            QMessageBox.warning(self, "提示", "请先登录")
            return
        pdf_name = item.text()
        if "  " in pdf_name:
            pdf_name = pdf_name.split("  ")[0].strip()
        self._current_selected_pdf = pdf_name
        if self.top_bar.exam_toggle.isChecked():
            self.load_quiz_for_pdf(pdf_name)
        else:
            self.open_pdf_by_name(pdf_name, 1, "")

    # ---------- 版本和更新（新方案） ----------
    def update_version_label(self):
        version = DATA_LOADER.get_current_version()
        self.version_label.setText(f"Version: {version}" if version else "Version: 开发模式")

    def refresh_after_sync(self):
        """
        在云同步成功后调用，强制 DataLoader 重新加载数据，并刷新所有界面组件。
        """
        # 强制 DataLoader 重新从文件加载（因为可能已更新 assets 目录）
        DATA_LOADER._load()   # 重新加载资产

        # 重新获取所有数据
        self.knowledge_data = DATA_LOADER.get_knowledge_graph()
        self.station_to_pdfs = DATA_LOADER.get_station_to_pdfs()
        self.file_categories = DATA_LOADER.get_file_categories()
        self.all_stations = DATA_LOADER.get_all_stations()

        docs = self.knowledge_data.get("documents", [])
        self.pdf_to_doc = {doc.get("pdf_name", ""): True for doc in docs if doc.get("pdf_name")}

        # 更新工站下拉框
        self.top_bar.set_stations(self.all_stations)
        if self.all_stations:
            self.current_station = self.all_stations[0]
        else:
            self.current_station = "无工站"

        # 刷新界面
        self.refresh_left_list()
        self.update_right_panel()
        self.update_version_label()
        self.statusBar().showMessage("数据已刷新")

    # ---------- 定时与手动同步槽函数 ----------
    def _on_silent_sync(self):
        """每小时静默同步（无弹窗，不打扰用户）"""
        self.update_checker.check_data_update(silent=True)

    def _on_refresh_clicked(self):
        """手动点击刷新按钮：执行完整同步（含进度弹窗），并显示提示"""
        self.update_checker.check_data_update(show_no_update=True, silent=False)

    # ---------- 图标 ----------
    def create_icon(self):
        from PyQt5.QtCore import QRect, QSize, Qt
        size = 256
        pixmap = QPixmap(QSize(size, size))
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setRenderHint(QPainter.TextAntialiasing)
        gradient = QRadialGradient(size/2, size/2, size/2)
        gradient.setColorAt(0, QColor(52, 152, 219))
        gradient.setColorAt(1, QColor(26, 82, 118))
        painter.setBrush(QBrush(gradient))
        painter.setPen(Qt.NoPen)
        painter.drawRoundedRect(0, 0, size, size, 20, 20)
        painter.setPen(Qt.NoPen)
        shadow_color = QColor(0, 0, 0, 80)
        painter.setBrush(QBrush(shadow_color))
        painter.drawRoundedRect(6, 8, size-12, size-12, 18, 18)
        painter.setBrush(QBrush(QColor(255, 255, 255, 230)))
        painter.drawRoundedRect(4, 4, size-8, size-8, 16, 16)
        font = QFont("Arial", 48, QFont.Bold)
        font.setStyleHint(QFont.SansSerif)
        painter.setFont(font)
        painter.setPen(QPen(QColor(26, 82, 118), 2))
        rect = QRect(0, 0, size, size)
        painter.drawText(rect, Qt.AlignCenter, "SOP")
        painter.setPen(QPen(QColor(52, 152, 219, 180), 3))
        painter.drawLine(30, 90, 90, 30)
        painter.end()
        return QIcon(pixmap)